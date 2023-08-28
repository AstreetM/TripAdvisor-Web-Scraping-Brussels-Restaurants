#imports here
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from bs4 import BeautifulSoup
import os
import wget
import time
from selenium.webdriver.chrome.service import Service
import pandas as pd
import xlrd
import requests
from urllib.request import urlopen
import urllib
import openpyxl
from selenium.webdriver.support.color import Color
import numpy as np
import csv
import openpyxl


headers = {
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'GET',
        'Access-Control-Allow-Headers': 'Content-Type',
        'accept': '*/*',
        'accept-encoding': 'gzip, deflate',
        'accept-language': 'en,mr;q=0.9',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36'}



# Est la pour chopper l'url et chopper la feuille excel avec les datas

wb1 = openpyxl.load_workbook("url_brussels_tripadvisor.xlsx")
sheet1 = wb1.active



list_url = []
for row in range(2,sheet1.max_row):
    list_url.append(sheet1.cell(row,1).value)
print(list_url)


print(list_url)
print(len(list_url))


list_de_url = []
list_noms = []
liste_emails =[]
listes_numero_tel = []
liste_adresse_complet = []
liste_site_web = []
list_ranking = []
liste_notes = []
listes_nb_avis = []



def obtenir_nom_rest(req):
    
    soup = BeautifulSoup(req.text, 'html.parser')
    soup.prettify()
    try:
        noms_resto = soup.find("h1")
    except:
        pass
    return noms_resto.get_text(strip=True)


def obtenir_email(req):
    # Obtenir l'E-mail
    
    soup = BeautifulSoup(req.text, 'html.parser')
    soup.prettify()
    list_emails = []
    try:
        for email in soup.select('a[href^=mailto]'):
            data = email['href']
            data = data.split('?')[0]
            data = data.replace('mailto:', '')
            list_emails.append(data)
        # Les listes vides sont évaluées à False
        if not list_emails:
            list_emails.append("Vide")
    except:
        liste_emails.append("Non disponible")
    return list_emails[0]




def obtenir_num_tel(req):
    
    soup = BeautifulSoup(req.text, 'html.parser')
    soup.prettify()
    num_list = []
    try:
        for num in soup.select('a[href^=tel]'):
            datas = num['href']
            datas = datas.replace('tel:', '')
            num_list.append(datas)
            if num_list[0] == "":
                num_list.append("NON disponible")
            if len(num_list) > 0:
                break
        # Les listes vides sont évaluées à False
        if not num_list:
            num_list.append("Vide")
    except:
        num_list.append("NON disponible")
    return num_list[0]


def obtenir_adresse(req):
        # Adresse
    soup = BeautifulSoup(req.text, 'html.parser')
    soup.prettify()
    adresse_list = []
    try:
        adresse = soup.find("span", {"class": "yEWoV"})
        adresse_list.append(adresse.get_text(strip=True))
        if adresse == "":
            adresse_list.append("NON disponible")
    except:
        adresse_list.append("NON disponible")
    return adresse_list[0]


def obtenir_ranking(req):
    #obtenir ranking
    soup = BeautifulSoup(req.text, 'html.parser')
    soup.prettify()
    ranking_list = []
    try:
        for numero_rest_ranking in soup.body.find_all(class_="cNFlb"):
            ranking_list.append(numero_rest_ranking.get_text(strip=True))
                # Les listes vides sont évaluées à False
        if not ranking_list:
            ranking_list.append("Vide")
    except:
        ranking_list.append("non disponible")
    return ranking_list[0]



def obtenir_notes(req):
    # obtenir note et nb avis (rajoute 2 trucs (fouchette prix et type de cuisine)
    soup = BeautifulSoup(req.text, 'html.parser')
    soup.prettify()
    notes_liste = []
    try:
        for element in soup.body.find_all(class_="YDAvY R2 F1 e k"):
            notes = element.find("span",{"class":"ZDEqb"})
            notes_liste.append(notes.get_text(strip=True))
        if not ranking_list:
            notes_liste.append("Vide")
    except:
        notes_liste.append("non disponible")
    return notes_liste[0]



def obtenir_nb_avis(req):
    # obtenir note et nb avis (rajoute 2 trucs (fouchette prix et type de cuisine)
    soup = BeautifulSoup(req.text, 'html.parser')
    soup.prettify()
    nb_avis_liste = []
    try:
        for element in soup.body.find_all(class_="YDAvY R2 F1 e k"):
            nb_avis = element.find("a",{"class":"IcelI"})
            string = nb_avis.get_text(strip=True)
            characters = "avis"
            for x in range(len(characters)):
                string = string.replace(characters[x], "")
            #print(string)
            nb_avis_liste.append(string)
        if not nb_avis_liste:
            nb_avis_liste.append("Vide")
    except:
        nb_avis_liste.append("non disponible")
    return nb_avis_liste[0]



def obtenir_site_web(req):
    soup = BeautifulSoup(req.text, 'html.parser')
    soup.prettify()
    web_list = []
    try:
        for num in soup.select('a[href^=https]'):
            datas = num['href']
            web_list.append(datas)
            if web_list == "":
                web_list.append("NON disponible")
            if len(web_list) > 0:
                break
        # Les listes vides sont évaluées à False
        if not web_list:
            web_list.append("Vide")
    except:
        web_list.append("NON disponible")
    return web_list[0]



 for e in list_url[0:4200]:
    try:
        req = requests.get(e,headers=headers,timeout=15,verify=True)
    except:
        print("Timeout occurred")
        req = requests.get(e,headers=headers,timeout=15,verify=True)
        time.sleep(2)
        
    soup = BeautifulSoup(req.text, 'html.parser')
    print (req.status_code)
    list_de_url.append(e)
    
    list_noms.append(obtenir_nom_rest(req))

    liste_emails.append(obtenir_email(req))

    listes_numero_tel.append(obtenir_num_tel(req))
    
    liste_adresse_complet.append(obtenir_adresse(req))
    #obtenir_image(req)
    list_ranking.append(obtenir_ranking(req))
    
    liste_notes.append(obtenir_notes(req))
    
    listes_nb_avis.append(obtenir_nb_avis(req))
    
print("c'est terminé")


print("url "+str(len(list_de_url)))
print("email "+str(len(liste_emails)))
print("numéro "+str(len(listes_numero_tel)))
print("Adresse "+str(len(liste_adresse_complet)))
print("Site web "+str(len(liste_site_web)))
print("Ranking "+str(len(list_ranking)))
print("Notes "+str(len(liste_notes)))
print("Nombre d'avis "+str(len(listes_nb_avis)))



dict_avis = {'url':list_de_url,'nom du restaurant' : list_noms,'email':liste_emails,'numéro de telephone':listes_numero_tel,'Adresse':liste_adresse_complet,"Ranking":list_ranking,"Notes":liste_notes,"Nombre d'avis":listes_nb_avis}
pd.DataFrame(dict_avis)

df = pd.DataFrame(dict_avis)

print(df)
df.to_excel(r'C:votre_chemin/python/resultats_new_tripadvisor.xlsx', index=False)